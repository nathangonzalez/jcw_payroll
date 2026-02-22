#!/usr/bin/env python3
"""
Labor regression plan: compare PDF (truth) vs DB and manual sources,
and compute suggested weekly hour deltas per employee.
"""

import argparse
import re
import sqlite3
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import fitz
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DATE_IN_NAME = re.compile(r"(\d{6})")
LINE_GROSS = re.compile(r"^\S+\s+([0-9]+\.?[0-9]*)\s+")
DAY_NAMES = {
    "mon",
    "monday",
    "tue",
    "tues",
    "tuesday",
    "wed",
    "wednesday",
    "thu",
    "thur",
    "thurs",
    "thursday",
    "fri",
    "friday",
    "sat",
    "saturday",
    "sun",
    "sunday",
}


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", required=True, help="Path to prod app.db")
    parser.add_argument("--pdfs", required=True, help="Comma-separated payroll PDF paths")
    parser.add_argument("--exports-root", required=True, help="Path to exports root")
    parser.add_argument("--weeks", default="", help="Comma-separated week folders (e.g. 'Week 2,Week 3')")
    parser.add_argument(
        "--manual-xlsx",
        default="",
        help="Optional comma-separated XLSX files with columns Date, Job, Task, Start, Lunch, End, Total.",
    )
    parser.add_argument(
        "--ocr-review",
        default="",
        help="Optional OCR review XLSX with Review sheet (date, employee, customer, hours).",
    )
    parser.add_argument("--dept", default="LABOR", help="Department label to include (default: LABOR)")
    parser.add_argument("--out", required=True, help="Output XLSX report")
    return parser.parse_args()


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def round_hours(value) -> float:
    try:
        return round(float(value), 2)
    except Exception:
        return 0.0


def parse_employee_from_filename(path: Path) -> str:
    name = path.stem
    name = re.sub(r"\s*\(\d+\)\s*$", "", name).strip()
    return name


def read_xls_rows(path: Path) -> List[List]:
    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_index(0)
    rows = []
    for r in range(sh.nrows):
        rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
    return rows


def read_xlsx_rows(path: Path) -> List[List]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def find_header_indices(rows: List[List]) -> Optional[Tuple[int, int, int, int]]:
    for idx, row in enumerate(rows[:10]):
        if not row:
            continue
        lowered = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        if "date" in lowered and "client name" in lowered and "hours per job" in lowered:
            return (idx, lowered.index("date"), lowered.index("client name"), lowered.index("hours per job"))
    return None


def is_day_name(value) -> bool:
    return isinstance(value, str) and value.strip().lower() in DAY_NAMES


def is_day_number(value) -> bool:
    return isinstance(value, (int, float)) and 1 <= int(value) <= 31


def build_date(year: int, month: int, day: int) -> str:
    return datetime(year, month, day).strftime("%Y-%m-%d")


def parse_manual_entries(path: Path, month_hint: Optional[str] = None) -> List[Tuple[str, str, str, float]]:
    rows = read_xlsx_rows(path) if path.suffix.lower() == ".xlsx" else read_xls_rows(path)
    header = find_header_indices(rows)
    if not header:
        return []
    header_row, date_col, client_col, hours_col = header
    file_mtime = datetime.utcfromtimestamp(path.stat().st_mtime)
    year = file_mtime.year
    month = file_mtime.month
    if month_hint:
        try:
            year, month = map(int, month_hint.split("-"))
        except Exception:
            pass

    entries = []
    pending = []
    current_date = None
    employee = parse_employee_from_filename(path)

    for row in rows[header_row + 1 :]:
        if not row or len(row) <= max(date_col, client_col, hours_col):
            continue
        date_cell = row[date_col]
        client_cell = row[client_col]
        hours_cell = row[hours_col]

        client = str(client_cell).strip() if client_cell is not None else ""
        hours = round_hours(hours_cell)

        if is_day_name(date_cell):
            if client and hours:
                pending.append((employee, client, hours))
            continue

        if is_day_number(date_cell):
            try:
                day_num = int(date_cell)
                use_year, use_month = year, month
                if not month_hint and day_num > file_mtime.day:
                    if use_month == 1:
                        use_year -= 1
                        use_month = 12
                    else:
                        use_month -= 1
                current_date = build_date(use_year, use_month, day_num)
            except ValueError:
                current_date = None
                pending = []
                continue
            if pending:
                for emp, cust, hrs in pending:
                    entries.append((current_date, emp, cust, hrs))
                pending = []
            if client and hours:
                entries.append((current_date, employee, client, hours))
            continue

        if current_date and client and hours:
            entries.append((current_date, employee, client, hours))

    return entries


def parse_voice_xlsx(path: Path) -> List[Tuple[str, str, str, float]]:
    wb = load_workbook(path, data_only=True)
    entries: List[Tuple[str, str, str, float]] = []
    for ws in wb.worksheets:
        employee = ws.title.replace("_", " ").strip()
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        try:
            date_idx = header.index("date")
            job_idx = header.index("job")
            total_idx = header.index("total")
        except ValueError:
            continue
        for row in rows[1:]:
            if not row or len(row) <= max(date_idx, job_idx, total_idx):
                continue
            date_val = row[date_idx]
            job_val = row[job_idx]
            total_val = row[total_idx]
            if not date_val or not total_val:
                continue
            date_str = str(date_val).strip()
            if len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
                date = date_str
            else:
                try:
                    date = datetime.strptime(date_str, "%m/%d/%Y").strftime("%Y-%m-%d")
                except Exception:
                    continue
            customer = str(job_val or "").strip() or "Unknown"
            hours = round_hours(total_val)
            if hours <= 0:
                continue
            entries.append((date, employee, customer, hours))
    return entries


def parse_ocr_review(path: Path) -> List[Tuple[str, str, str, float]]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    if "Review" not in wb.sheetnames:
        return []
    ws = wb["Review"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    try:
        date_idx = header.index("date")
        emp_idx = header.index("employee")
        cust_idx = header.index("customer")
        hours_idx = header.index("hours")
    except ValueError:
        return []
    entries: List[Tuple[str, str, str, float]] = []
    for row in rows[1:]:
        if not row or len(row) <= max(date_idx, emp_idx, cust_idx, hours_idx):
            continue
        date_val = row[date_idx]
        emp_val = row[emp_idx]
        cust_val = row[cust_idx]
        hours_val = row[hours_idx]
        if not date_val or not emp_val or not hours_val:
            continue
        date_str = str(date_val).strip()
        if len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
            date = date_str
        else:
            try:
                date = datetime.strptime(date_str, "%m/%d/%Y").strftime("%Y-%m-%d")
            except Exception:
                continue
        employee = str(emp_val).strip()
        customer = str(cust_val or "Unknown").strip() or "Unknown"
        hours = round_hours(hours_val)
        entries.append((date, employee, customer, hours))
    return entries


def parse_employee_name(raw: str) -> str:
    raw = raw.strip()
    if "," in raw:
        last, first = [p.strip() for p in raw.split(",", 1)]
        return f"{first} {last}".strip()
    return raw


def extract_pdf_gross(path: Path, dept_filter: str) -> Dict[str, float]:
    doc = fitz.open(path)
    gross_by_emp: Dict[str, float] = defaultdict(float)
    pending_gross = None
    current_dept = None
    dept_filter = (dept_filter or "").strip().upper()
    for page in doc:
        text = page.get_text("text")
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            if line.startswith("* "):
                current_dept = line.upper()
                continue
            if line.startswith("** Total"):
                current_dept = None
            match = LINE_GROSS.match(line)
            if match:
                try:
                    pending_gross = float(match.group(1))
                except Exception:
                    pending_gross = None
                continue
            if pending_gross is not None and "," in line:
                if not dept_filter or (current_dept and dept_filter in current_dept):
                    emp = parse_employee_name(line)
                    gross_by_emp[normalize_text(emp)] += pending_gross
                pending_gross = None
    return gross_by_emp


def week_range_from_filename(path: Path) -> Tuple[str, str]:
    match = DATE_IN_NAME.search(path.name)
    if not match:
        raise ValueError(f"Missing date in filename: {path.name}")
    mmddyy = match.group(1)
    week_end = datetime.strptime(mmddyy, "%m%d%y")
    week_start = week_end - timedelta(days=6)
    return week_start.strftime("%Y-%m-%d"), week_end.strftime("%Y-%m-%d")


def week_key_for_date(date_str: str, week_ranges: List[Tuple[str, str, str]]) -> Optional[str]:
    for week_id, start, end in week_ranges:
        if start <= date_str <= end:
            return week_id
    return None


def load_db_rates(conn) -> Dict[str, float]:
    cur = conn.cursor()
    cur.execute("SELECT name, default_pay_rate FROM employees")
    return {normalize_text(r[0]): float(r[1] or 0) for r in cur.fetchall()}


def db_hours_for_week(conn, week_start: str, week_end: str) -> Dict[str, float]:
    cur = conn.cursor()
    cur.execute(
        """
        SELECT e.name, te.hours
        FROM time_entries te
        JOIN employees e ON e.id = te.employee_id
        WHERE te.work_date >= ? AND te.work_date <= ?
        """,
        (week_start, week_end),
    )
    hours_by_emp: Dict[str, float] = defaultdict(float)
    for name, hours in cur.fetchall():
        hours_by_emp[normalize_text(name)] += float(hours or 0)
    return hours_by_emp


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(40, max(12, max_len + 2))


def main():
    args = parse_args()
    db_path = Path(args.db)
    exports_root = Path(args.exports_root)
    pdf_paths = [Path(p.strip()) for p in args.pdfs.split(",") if p.strip()]

    week_dirs = []
    if args.weeks:
        for name in [w.strip() for w in args.weeks.split(",") if w.strip()]:
            week_dirs.append(exports_root / name)
    else:
        week_dirs = sorted([p for p in exports_root.iterdir() if p.is_dir() and p.name.lower().startswith("week")])

    manual_entries: List[Tuple[str, str, str, float]] = []
    for week_dir in week_dirs:
        for path in sorted(week_dir.glob("*.xls*")):
            manual_entries.extend(parse_manual_entries(path))
    if args.manual_xlsx:
        for path in [p.strip() for p in args.manual_xlsx.split(",") if p.strip()]:
            manual_entries.extend(parse_voice_xlsx(Path(path)))
    if args.ocr_review:
        manual_entries.extend(parse_ocr_review(Path(args.ocr_review)))

    conn = sqlite3.connect(db_path)
    rates = load_db_rates(conn)

    # Prepare week ranges + pdf gross
    week_ranges: List[Tuple[str, str, str]] = []
    pdf_gross_by_week: Dict[str, Dict[str, float]] = {}
    for pdf in pdf_paths:
        week_start, week_end = week_range_from_filename(pdf)
        week_id = week_end
        week_ranges.append((week_id, week_start, week_end))
        pdf_gross_by_week[week_id] = extract_pdf_gross(pdf, args.dept)

    # Manual hours by week/employee
    manual_hours_by_week: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for date, employee, _, hours in manual_entries:
        week_id = week_key_for_date(date, week_ranges)
        if not week_id:
            continue
        manual_hours_by_week[week_id][normalize_text(employee)] += float(hours or 0)

    wb = Workbook()
    wb.remove(wb.active)
    green = PatternFill("solid", fgColor="D6F5D6")
    red = PatternFill("solid", fgColor="FFD6D6")
    yellow = PatternFill("solid", fgColor="FFF2CC")

    summary = wb.create_sheet("Summary")
    summary.append(
        [
            "week_end",
            "week_start",
            "pdf_gross_total",
            "db_gross_total",
            "diff_gross",
            "manual_hours_total",
            "db_hours_total",
            "diff_hours",
        ]
    )
    for cell in summary[1]:
        cell.font = Font(bold=True)

    plan = wb.create_sheet("Plan")
    plan.append(
        [
            "week_end",
            "employee",
            "pdf_gross",
            "db_hours",
            "db_rate",
            "db_gross",
            "manual_hours",
            "target_hours",
            "delta_hours",
            "delta_gross",
            "status",
        ]
    )
    for cell in plan[1]:
        cell.font = Font(bold=True)

    for week_id, week_start, week_end in sorted(week_ranges):
        pdf_gross = pdf_gross_by_week.get(week_id, {})
        db_hours = db_hours_for_week(conn, week_start, week_end)
        manual_hours = manual_hours_by_week.get(week_id, {})

        all_emps = set(pdf_gross) | set(db_hours) | set(manual_hours)
        pdf_total = 0.0
        db_total = 0.0
        manual_total = 0.0
        db_hours_total = 0.0

        for emp in sorted(all_emps):
            pdf_val = round(float(pdf_gross.get(emp, 0)), 2)
            hours_val = round(float(db_hours.get(emp, 0)), 2)
            manual_val = round(float(manual_hours.get(emp, 0)), 2)
            rate = round(float(rates.get(emp, 0)), 2)
            db_gross = round(hours_val * rate, 2)
            diff_gross = round(pdf_val - db_gross, 2)

            if manual_val > 0:
                target_hours = manual_val
                status = "manual_target"
            elif pdf_val > 0 and rate > 0:
                target_hours = round(pdf_val / rate, 2)
                status = "estimate_from_pdf"
            else:
                target_hours = ""
                status = "missing_rate" if pdf_val > 0 else "no_pdf"

            delta_hours = round(target_hours - hours_val, 2) if isinstance(target_hours, (int, float)) else ""
            delta_gross = round((target_hours * rate) - db_gross, 2) if isinstance(target_hours, (int, float)) else ""

            pdf_total += pdf_val
            db_total += db_gross
            manual_total += manual_val
            db_hours_total += hours_val

            plan.append(
                [
                    week_end,
                    emp,
                    pdf_val,
                    hours_val,
                    rate,
                    db_gross,
                    manual_val,
                    target_hours,
                    delta_hours,
                    delta_gross,
                    status,
                ]
            )
            row = plan[plan.max_row]
            if status == "manual_target":
                for cell in row:
                    cell.fill = green
            elif status == "estimate_from_pdf":
                for cell in row:
                    cell.fill = yellow
            elif status == "missing_rate":
                for cell in row:
                    cell.fill = red

        summary.append(
            [
                week_end,
                week_start,
                round(pdf_total, 2),
                round(db_total, 2),
                round(pdf_total - db_total, 2),
                round(manual_total, 2),
                round(db_hours_total, 2),
                round(manual_total - db_hours_total, 2),
            ]
        )

    autosize(plan)
    autosize(summary)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved {out_path}")


if __name__ == "__main__":
    main()
