"""
Final Labor Reconciliation Report
Compares: PDF (truth) vs Prod Export vs Manual Timesheets
Outputs: Highlighted XLSX showing exactly what needs to change in prod
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── COLORS ──
RED = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
GREEN = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
ORANGE = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
BLUE_HDR = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
GRAY_HDR = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True, size=11)
BOLD_RED = Font(bold=True, size=11, color="CC0000")
MONEY = '#,##0.00'
HRS = '0.000'
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# ── PDF TRUTH (LABOR dept only) ─────────────────────────────────
# Derived from the 3 labor distribution PDFs
PDF = {
    "Week 1": {  # ending 2/3/26 - Batch 738
        "Boban Abbate":   {"gross": 1700.00, "rate": 42.50, "hrs": 40.0},
        "Thomas Brinson": {"gross": 1400.00, "rate": 35.00, "hrs": 40.0},
        "Jason Green":    {"gross": 1400.00, "rate": 35.00, "hrs": 40.0},
        "Phil Henderson": {"gross": 1245.00, "rate": 30.00, "hrs": 41.5},
        "Doug Kinsey":    {"gross": 1278.75, "rate": 30.00, "hrs": 42.625},
        "Sean Matthew":   {"gross":  800.00, "rate": 20.00, "hrs": 40.0},
    },
    "Week 2": {  # ending 2/10/26 - Batch 740
        "Boban Abbate":   {"gross": 1700.00, "rate": 42.50, "hrs": 40.0},
        "Thomas Brinson": {"gross": 1400.00, "rate": 35.00, "hrs": 40.0},
        "Jason Green":    {"gross": 1400.00, "rate": 35.00, "hrs": 40.0},
        "Phil Henderson": {"gross": 1200.00, "rate": 30.00, "hrs": 40.0},
        "Doug Kinsey":    {"gross": 1211.25, "rate": 30.00, "hrs": 40.375},
        "Sean Matthew":   {"gross":  800.00, "rate": 20.00, "hrs": 40.0},
    },
    "Week 3": {  # ending 2/17/26 - Batch 742
        "Boban Abbate":   {"gross": 1700.00, "rate": 42.50, "hrs": 40.0},
        "Thomas Brinson": {"gross": 1365.00, "rate": 35.00, "hrs": 39.0},
        "Jason Green":    {"gross": 1400.00, "rate": 35.00, "hrs": 40.0},
        "Phil Henderson": {"gross": 1335.00, "rate": 30.00, "hrs": 44.5},
        "Doug Kinsey":    {"gross": 1402.50, "rate": 30.00, "hrs": 46.75},
        "Sean Matthew":   {"gross":  830.00, "rate": 20.00, "hrs": 41.5},
    },
}

# ── PROD EXPORT (parsed from Payroll_Breakdown xlsx) ────────────
# Hours computed from column F summing in each employee sheet block
PROD = {
    "Week 1": {
        "Boban Abbate":   {"hrs": 40.0, "entries": "Behrens 6, Boyle 32, Theobald 1.5, Walsh 0.5"},
        "Thomas Brinson": {"hrs": 40.0, "entries": "Boyle 8.5, Landy 31.5"},
        "Jason Green":    {"hrs": 40.0, "entries": "Boyle 10, Gee 0.5, Landy 27, Lucas 1, O'Connor 0.5, Schroeder 1"},
        "Phil Henderson": {"hrs": 41.0, "entries": "Watkins 41 (5 days: 8,8,9,8,8)"},
        "Doug Kinsey":    {"hrs": 41.75, "entries": "Boyle 19.5, Landy 6.5, Lynn 5.5, PTO 8, Watkins 2.25"},
        "Sean Matthew":   {"hrs": 40.0, "entries": "Boyle 25, Landy 11.5, Watkins 3.5"},
    },
    "Week 2": {
        "Boban Abbate":   {"hrs": 39.0, "entries": "Boyle 36, Campbell 1, Sweeney 0.5, Walsh 0.5, Walsh-Maint 0.5 + missing 0.5"},
        "Thomas Brinson": {"hrs": 36.5, "entries": "Boyle 8.5, Landy 28 (some reconciled from manual)"},
        "Jason Green":    {"hrs": 37.0, "entries": "Boyle 8, Landers 0.5, Landy 4.5, Lucas 2.5, McFarland 6.5, Muncey 4, Richer 3.5, Schauer 1, Schroeder 1, Tubergen 3, Turbergen 1 + missing"},
        "Phil Henderson": {"hrs": 38.5, "entries": "Tubergen 0.5+0.5, Watkins 37.5 (some reconciled)"},
        "Doug Kinsey":    {"hrs": 40.25, "entries": "Boyle 10.5, JCW 15, Lynn 2, PTO 8, Watkins 4.75"},
        "Sean Matthew":   {"hrs": 39.0, "entries": "Boyle 23, Office 8, PTO 8"},
    },
    "Week 3": {
        "Boban Abbate":   {"hrs": 0, "entries": "NOT IN PROD"},
        "Thomas Brinson": {"hrs": 2.5, "entries": "Landy 2.5 only"},
        "Jason Green":    {"hrs": 0, "entries": "NOT IN PROD"},
        "Phil Henderson": {"hrs": 27.0, "entries": "Watkins 27 (3 days of 8 + 1 Sat 3)"},
        "Doug Kinsey":    {"hrs": 35.0, "entries": "Boyle 16, Gonzalez 10.5, Howard 1, Jebsen 3.25, Lynn 0.5, Watkins 2.5, Welles 0.5 + partial"},
        "Sean Matthew":   {"hrs": 0, "entries": "NOT IN PROD"},
    },
}

# ── MANUAL TIMESHEETS ──────────────────────────────────────────
# Week 1: from voice-to-text/OCR (exports/week 1 photos/)
# Week 2: from 2_4/ folder XLS files
# Week 3: from exports/Week 3/ folder XLS files
MANUAL = {
    "Week 1": {  # Voice-to-text source
        "Boban Abbate":   {"hrs": 40.0, "clients": "Behrens 6, Boyle 32, Theobald 1.5, Walsh 0.5", "source": "voice/OCR"},
        "Thomas Brinson": {"hrs": 40.0, "clients": "Boyle 8.5, Landy 31.5", "source": "voice/OCR"},
        "Jason Green":    {"hrs": 40.0, "clients": "Boyle 10, Gee 0.5, Landy 27, Lucas 1, O'Connor 0.5, Schroeder 1", "source": "voice/OCR - email says 39.5 total but PDF=40"},
        "Phil Henderson": {"hrs": 41.5, "clients": "Watkins 41.5", "source": "voice/OCR - handwritten sheet hard to read"},
        "Doug Kinsey":    {"hrs": 42.625, "clients": "Boyle 19.5, Landy 6.5, Lynn 5.5, PTO 8, Watkins 2.25 + 0.875 unaccounted", "source": "voice/OCR"},
        "Sean Matthew":   {"hrs": 40.0, "clients": "Boyle 25, Landy 11.5, Watkins 3.5", "source": "voice/OCR"},
    },
    "Week 2": {  # from 2_4/ XLS files
        "Boban Abbate":   {"hrs": 40.0, "clients": "Boyle 36, Campbell 1, Sweeney 1, Walsh-Insp 1, Walsh-Maint 1", "source": "2_4/Boban Abatte (3).xls"},
        "Thomas Brinson": {"hrs": 40.0, "clients": "Boyle 17.5, Delacruz NB 0.5, Landy 21.5, PTO 0.5", "source": "2_4/Thomas Brinson (2).xls"},
        "Jason Green":    {"hrs": 40.0, "clients": "Boyle 7.5, Jebsen 1, Landers 0.5, Landy 4.5, Lucas 2.5, McFarland 6, Muncey 3.5, Richer 3.5, Schauer 1, Schroeder 1, Tubergen 7, Watkins 1, Salary 1", "source": "2_4/Jason Green (5).xls"},
        "Phil Henderson": {"hrs": 40.0, "clients": "Tubergen 1.5, Watkins 38.5", "source": "2_4/Phil Henderson (2).xls"},
        "Doug Kinsey":    {"hrs": 40.125, "clients": "Boyle 10.5, Lynn 2, Watkins 4.75, JCW Shop 15, PTO 8, OT calc 0.125", "source": "2_4/Doug Kinsey (2).xls"},
        "Sean Matthew":   {"hrs": 40.0, "clients": "Boyle 24, JCW Shop 8, PTO 8", "source": "2_4/Sean Matthew (2).xls"},
    },
    "Week 3": {  # from exports/Week 3/ XLS files
        "Boban Abbate":   {"hrs": 40.0, "clients": "Boyle 38, Howard 0.5, Sweeney 0.5, Walsh-Maint 1", "source": "Week 3/Boban Abatte (4).xls"},
        "Thomas Brinson": {"hrs": 39.0, "clients": "Boyle 8, Gonzalez 9.5, Jebsen 3, Landy 18.5", "source": "Week 3/Thomas Brinson (3).xls"},
        "Jason Green":    {"hrs": 40.0, "clients": "Boyle 29.5, Howard 1.5, Landy 2.5, Lucas 0.5, Muncey 1.5, Schroeder 2.5, Salary-LeftEarly 2", "source": "Week 3/Jason Green (6).xls"},
        "Phil Henderson": {"hrs": 44.5, "clients": "Watkins 43, OT calc 1.5", "source": "Week 3/Phil Henderson (3).xls"},
        "Doug Kinsey":    {"hrs": 46.75, "clients": "Boyle 17.75, Gonzalez 12, Hall 2.25, Howard 1, Jebsen 7, Lynn 0.5, Watkins 3.25, Welles 0.75, OT calc 2.25", "source": "Week 3/Doug Kinsey (3).xls"},
        "Sean Matthew":   {"hrs": 41.5, "clients": "Boyle 26.5, Gonzalez 9, Hall 2.25, Jebsen 3.25, OT calc 0.5", "source": "Week 3/Sean Matthew (3).xls"},
    },
}

EMPLOYEES = ["Boban Abbate", "Thomas Brinson", "Jason Green", "Phil Henderson", "Doug Kinsey", "Sean Matthew"]
WEEKS = ["Week 1", "Week 2", "Week 3"]
RATES = {e: PDF["Week 1"][e]["rate"] for e in EMPLOYEES}


def hdr_cell(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = BLUE_HDR
    c.font = WHITE_FONT
    c.border = BORDER
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    return c

def data_cell(ws, row, col, val, fmt=None, fill=None, font=None):
    c = ws.cell(row=row, column=col, value=val)
    c.border = BORDER
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if font: c.font = font
    return c


def main():
    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════
    # SHEET 1: Executive Summary
    # ════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    
    # Title
    ws.merge_cells('A1:T1')
    c = ws.cell(row=1, column=1, value="LABOR RECONCILIATION — Feb 2026 Weeks 1-3")
    c.font = Font(bold=True, size=14)
    
    ws.merge_cells('A2:T2')
    ws.cell(row=2, column=1, value="PDF Labor Report = Source of Truth  |  Yellow = needs adjustment  |  Red = missing from prod  |  Green = matches").font = Font(italic=True, size=10)

    # Headers
    row = 4
    headers = ["Employee", "$/hr"]
    for wk in WEEKS:
        headers += [f"{wk}\nPDF Hrs", f"{wk}\nProd Hrs", f"{wk}\nManual Hrs", f"{wk}\nΔ Hrs", f"{wk}\nPDF $", f"{wk}\nΔ $"]
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, row, ci, h)

    # Data
    row = 5
    total_gap = 0
    for emp in EMPLOYEES:
        rate = RATES[emp]
        data_cell(ws, row, 1, emp, font=BOLD)
        data_cell(ws, row, 2, rate, fmt=MONEY)
        
        for wi, wk in enumerate(WEEKS):
            base_col = 3 + wi * 6
            
            pdf_hrs = PDF[wk][emp]["hrs"]
            pdf_gross = PDF[wk][emp]["gross"]
            prod_hrs = PROD[wk][emp]["hrs"]
            manual_hrs = MANUAL[wk][emp]["hrs"]
            delta_hrs = prod_hrs - pdf_hrs
            delta_gross = delta_hrs * rate
            total_gap += abs(delta_gross)
            
            match = abs(delta_hrs) < 0.01
            
            data_cell(ws, row, base_col, pdf_hrs, fmt='0.000')
            data_cell(ws, row, base_col+1, prod_hrs, fmt='0.000',
                      fill=GREEN if match else (RED if prod_hrs == 0 else YELLOW))
            data_cell(ws, row, base_col+2, manual_hrs, fmt='0.000')
            data_cell(ws, row, base_col+3, delta_hrs, fmt='+0.000;-0.000;0',
                      fill=GREEN if match else RED, 
                      font=BOLD_RED if not match else None)
            data_cell(ws, row, base_col+4, pdf_gross, fmt=MONEY)
            data_cell(ws, row, base_col+5, delta_gross, fmt='+#,##0.00;-#,##0.00;$0',
                      fill=GREEN if match else RED,
                      font=BOLD_RED if not match else None)
        row += 1

    # Totals row
    row += 1
    data_cell(ws, row, 1, "TOTALS", font=BOLD)
    for wi, wk in enumerate(WEEKS):
        base_col = 3 + wi * 6
        pdf_total = sum(PDF[wk][e]["gross"] for e in EMPLOYEES)
        prod_total = sum(PROD[wk][e]["hrs"] * RATES[e] for e in EMPLOYEES)
        manual_total = sum(MANUAL[wk][e]["hrs"] * RATES[e] for e in EMPLOYEES)
        delta = prod_total - pdf_total
        match = abs(delta) < 0.01
        
        data_cell(ws, row, base_col+4, pdf_total, fmt=MONEY, font=BOLD)
        data_cell(ws, row, base_col+5, delta, fmt='+#,##0.00;-#,##0.00;$0',
                  fill=GREEN if match else RED, font=BOLD)

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 7
    for c in range(3, 21):
        ws.column_dimensions[get_column_letter(c)].width = 11

    # ════════════════════════════════════════════════════════════════
    # SHEET 2: Fixes Needed (actionable)
    # ════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Fixes Needed")
    
    ws2.merge_cells('A1:H1')
    ws2.cell(row=1, column=1, value="PROD EXPORT — CELLS/FORMULAS TO ADJUST").font = Font(bold=True, size=13)
    
    fix_headers = ["Employee", "Week", "Prod Sheet / Location", 
                   "Current Prod Hrs", "Should Be (PDF)", "Δ Hours",
                   "Action", "Manual Source Reference"]
    for ci, h in enumerate(fix_headers, 1):
        hdr_cell(ws2, 3, ci, h)
    
    frow = 4
    for wk in WEEKS:
        for emp in EMPLOYEES:
            pdf_hrs = PDF[wk][emp]["hrs"]
            prod_hrs = PROD[wk][emp]["hrs"]
            manual_hrs = MANUAL[wk][emp]["hrs"]
            delta = prod_hrs - pdf_hrs
            
            if abs(delta) < 0.01:
                continue
            
            # Determine which sheet/block in the prod export
            if wk == "Week 1":
                if emp in ["Doug Kinsey", "Phil Henderson", "Thomas Brinson"]:
                    sheet_loc = f"'{emp}' sheet, 3rd block (rows ~89-124)"
                else:
                    sheet_loc = f"'{emp}' sheet, 2nd block (rows ~46-81)"
            elif wk == "Week 2":
                if emp in ["Doug Kinsey", "Phil Henderson", "Thomas Brinson"]:
                    sheet_loc = f"'{emp}' sheet, 2nd block (rows ~46-81)"
                else:
                    sheet_loc = f"'{emp}' sheet, 1st block (rows ~3-38)"
            else:  # Week 3
                if emp in ["Doug Kinsey", "Phil Henderson", "Thomas Brinson"]:
                    sheet_loc = f"'{emp}' sheet, 1st block (rows ~3-38)"
                else:
                    sheet_loc = f"NOT IN PROD — need to add week block"
                    
            if prod_hrs == 0:
                action = f"ADD entire week: {pdf_hrs}h from manual"
            else:
                action = f"ADD {abs(delta):.3f}h (adjust entries to total {pdf_hrs}h)"
            
            manual_ref = MANUAL[wk][emp]["source"]
            manual_clients = MANUAL[wk][emp]["clients"]
            
            data_cell(ws2, frow, 1, emp, font=BOLD)
            data_cell(ws2, frow, 2, wk)
            data_cell(ws2, frow, 3, sheet_loc, fill=RED if prod_hrs == 0 else YELLOW)
            data_cell(ws2, frow, 4, prod_hrs, fmt='0.000')
            data_cell(ws2, frow, 5, pdf_hrs, fmt='0.000', font=BOLD)
            data_cell(ws2, frow, 6, delta, fmt='+0.000;-0.000;0', fill=RED, font=BOLD_RED)
            data_cell(ws2, frow, 7, action, fill=ORANGE)
            data_cell(ws2, frow, 8, f"{manual_ref}\n{manual_clients}")
            ws2.cell(row=frow, column=8).alignment = Alignment(wrap_text=True)
            
            frow += 1
    
    # Summary of total gap
    frow += 1
    data_cell(ws2, frow, 1, "TOTAL GAP:", font=BOLD)
    total_gap_dollars = sum(
        (PROD[wk][e]["hrs"] - PDF[wk][e]["hrs"]) * RATES[e]
        for wk in WEEKS for e in EMPLOYEES
        if PROD[wk][e]["hrs"] - PDF[wk][e]["hrs"] < -0.01
    )
    data_cell(ws2, frow, 6, total_gap_dollars, fmt=MONEY, fill=RED, font=BOLD_RED)
    data_cell(ws2, frow, 7, "Total $ missing from prod vs PDF truth", font=BOLD)
    
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 40
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 10
    ws2.column_dimensions['G'].width = 40
    ws2.column_dimensions['H'].width = 60

    # ════════════════════════════════════════════════════════════════
    # SHEET 3: Per-Employee Client Breakdown (Manual vs Prod)
    # ════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Client Breakdown")
    
    ws3.merge_cells('A1:F1')
    ws3.cell(row=1, column=1, value="PER-EMPLOYEE CLIENT HOURS — Manual Timesheet vs Prod Export").font = Font(bold=True, size=13)
    
    det_headers = ["Employee", "Week", "Source", "Client Breakdown", "Total Hours", "PDF Hours"]
    for ci, h in enumerate(det_headers, 1):
        hdr_cell(ws3, 3, ci, h)
    
    drow = 4
    for emp in EMPLOYEES:
        for wk in WEEKS:
            pdf_hrs = PDF[wk][emp]["hrs"]
            prod_hrs = PROD[wk][emp]["hrs"]
            manual_hrs = MANUAL[wk][emp]["hrs"]
            
            match_prod = abs(prod_hrs - pdf_hrs) < 0.01
            match_manual = abs(manual_hrs - pdf_hrs) < 0.01
            
            # Manual row
            data_cell(ws3, drow, 1, emp, font=BOLD)
            data_cell(ws3, drow, 2, wk)
            data_cell(ws3, drow, 3, "MANUAL", fill=GRAY_HDR)
            data_cell(ws3, drow, 4, MANUAL[wk][emp]["clients"])
            ws3.cell(row=drow, column=4).alignment = Alignment(wrap_text=True)
            data_cell(ws3, drow, 5, manual_hrs, fmt='0.000', fill=GREEN if match_manual else YELLOW)
            data_cell(ws3, drow, 6, pdf_hrs, fmt='0.000')
            drow += 1
            
            # Prod row
            data_cell(ws3, drow, 1, "")
            data_cell(ws3, drow, 2, "")
            data_cell(ws3, drow, 3, "PROD", fill=GRAY_HDR)
            data_cell(ws3, drow, 4, PROD[wk][emp]["entries"])
            ws3.cell(row=drow, column=4).alignment = Alignment(wrap_text=True)
            data_cell(ws3, drow, 5, prod_hrs, fmt='0.000', 
                      fill=GREEN if match_prod else (RED if prod_hrs == 0 else YELLOW))
            data_cell(ws3, drow, 6, "")
            drow += 1
            drow += 1  # spacing
    
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 10
    ws3.column_dimensions['D'].width = 80
    ws3.column_dimensions['E'].width = 12
    ws3.column_dimensions['F'].width = 10

    # ════════════════════════════════════════════════════════════════
    # SAVE
    # ════════════════════════════════════════════════════════════════
    out_path = "exports/reconcile/2026-02/labor_reconciliation_final.xlsx"
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"✅ Saved: {out_path}")

    # ── Console Summary ───────────────────────────────────────────
    print("\n" + "=" * 80)
    print("LABOR RECONCILIATION — PROD vs PDF (Source of Truth)")
    print("=" * 80)
    
    for wk in WEEKS:
        pdf_total = sum(PDF[wk][e]["gross"] for e in EMPLOYEES)
        prod_total = sum(PROD[wk][e]["hrs"] * RATES[e] for e in EMPLOYEES)
        delta = prod_total - pdf_total
        status = "✅" if abs(delta) < 0.01 else "❌"
        print(f"\n{status} {wk} (PDF=${pdf_total:,.2f}, Prod=${prod_total:,.2f}, Δ=${delta:+,.2f})")
        
        for emp in EMPLOYEES:
            p = PDF[wk][emp]
            prod_hrs = PROD[wk][emp]["hrs"]
            dh = prod_hrs - p["hrs"]
            if abs(dh) < 0.01:
                print(f"   ✅ {emp}: {prod_hrs}h = PDF {p['hrs']}h")
            else:
                print(f"   ❌ {emp}: Prod={prod_hrs}h, PDF={p['hrs']}h, "
                      f"Δ={dh:+.3f}h (${dh * p['rate']:+.2f})")
                if prod_hrs == 0:
                    print(f"      ➡️  LOAD FROM: {MANUAL[wk][emp]['source']}")
                else:
                    print(f"      ➡️  Adjust +{abs(dh):.3f}h in prod")

    print(f"\n{'=' * 80}")
    total_missing = sum(
        (PDF[wk][e]["hrs"] - PROD[wk][e]["hrs"]) * RATES[e]
        for wk in WEEKS for e in EMPLOYEES
        if PROD[wk][e]["hrs"] < PDF[wk][e]["hrs"]
    )
    print(f"TOTAL MISSING FROM PROD: ${total_missing:,.2f}")
    print(f"{'=' * 80}")


if __name__ == "__main__":
    main()