#!/usr/bin/env python3
"""
Populate the Payroll Breakdown template with manual entries (exports + voice + OCR review),
preserving formulas. Highlights rows that need verification (OCR/voice/unknown customer).
"""

import argparse
import re
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Set

import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DAY_NAMES = {
    "mon",
    "monday",
    "tue",
    "tues",
    "tuesday",
    "wed",
    "wednesday",
    "thu",
    "thur",
    "thurs",
    "thursday",
    "fri",
    "friday",
    "sat",
    "saturday",
    "sun",
    "sunday",
}


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True, help="Path to Payroll_Breakdown_*.xlsx")
    parser.add_argument("--out", required=True, help="Output XLSX path")
    parser.add_argument("--exports-root", required=True, help="Path to exports root")
    parser.add_argument("--weeks", default="", help="Comma-separated week folders (e.g. 'Week 2,Week 3')")
    parser.add_argument(
        "--manual-xlsx",
        default="",
        help="Optional comma-separated XLSX files with columns Date, Job, Task, Start, Lunch, End, Total.",
    )
    parser.add_argument(
        "--ocr-review",
        default="",
        help="Optional OCR review XLSX with Review sheet (date, employee, customer, hours).",
    )
    parser.add_argument("--db", help="Optional path to SQLite DB for drift check")
    return parser.parse_args()


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def round_hours(value) -> float:
    try:
        return round(float(value), 2)
    except Exception:
        return 0.0


def parse_employee_from_filename(path: Path) -> str:
    name = path.stem
    name = re.sub(r"\s*\(\d+\)\s*$", "", name).strip()
    return name


def read_xls_rows(path: Path) -> List[List]:
    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_index(0)
    rows = []
    for r in range(sh.nrows):
        rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
    return rows


def read_xlsx_rows(path: Path) -> List[List]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def find_header_indices(rows: List[List]) -> Optional[Tuple[int, int, int, int]]:
    for idx, row in enumerate(rows[:10]):
        if not row:
            continue
        lowered = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        if "date" in lowered and "client name" in lowered and "hours per job" in lowered:
            return (idx, lowered.index("date"), lowered.index("client name"), lowered.index("hours per job"))
    return None


def is_day_name(value) -> bool:
    return isinstance(value, str) and value.strip().lower() in DAY_NAMES


def is_day_number(value) -> bool:
    return isinstance(value, (int, float)) and 1 <= int(value) <= 31


def build_date(year: int, month: int, day: int) -> str:
    return datetime(year, month, day).strftime("%Y-%m-%d")


def parse_manual_entries(path: Path, month_hint: Optional[str] = None) -> List[Tuple[str, str, str, float, str]]:
    rows = read_xlsx_rows(path) if path.suffix.lower() == ".xlsx" else read_xls_rows(path)
    header = find_header_indices(rows)
    if not header:
        return []
    header_row, date_col, client_col, hours_col = header
    file_mtime = datetime.utcfromtimestamp(path.stat().st_mtime)
    year = file_mtime.year
    month = file_mtime.month
    if month_hint:
        try:
            year, month = map(int, month_hint.split("-"))
        except Exception:
            pass

    entries = []
    pending = []
    current_date = None
    employee = parse_employee_from_filename(path)
    source = f"export:{path.name}"

    for row in rows[header_row + 1 :]:
        if not row or len(row) <= max(date_col, client_col, hours_col):
            continue
        date_cell = row[date_col]
        client_cell = row[client_col]
        hours_cell = row[hours_col]

        client = str(client_cell).strip() if client_cell is not None else ""
        hours = round_hours(hours_cell)

        if is_day_name(date_cell):
            if client and hours:
                pending.append((employee, client, hours))
            continue

        if is_day_number(date_cell):
            try:
                day_num = int(date_cell)
                use_year, use_month = year, month
                if not month_hint and day_num > file_mtime.day:
                    if use_month == 1:
                        use_year -= 1
                        use_month = 12
                    else:
                        use_month -= 1
                current_date = build_date(use_year, use_month, day_num)
            except ValueError:
                current_date = None
                pending = []
                continue
            if pending:
                for emp, cust, hrs in pending:
                    entries.append((current_date, emp, cust, hrs, source))
                pending = []
            if client and hours:
                entries.append((current_date, employee, client, hours, source))
            continue

        if current_date and client and hours:
            entries.append((current_date, employee, client, hours, source))

    return entries


def parse_voice_xlsx(path: Path) -> List[Tuple[str, str, str, float, str]]:
    wb = load_workbook(path, data_only=True)
    entries: List[Tuple[str, str, str, float, str]] = []
    for ws in wb.worksheets:
        employee = ws.title.replace("_", " ").strip()
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        try:
            date_idx = header.index("date")
            job_idx = header.index("job")
            total_idx = header.index("total")
        except ValueError:
            continue
        for row in rows[1:]:
            if not row or len(row) <= max(date_idx, job_idx, total_idx):
                continue
            date_val = row[date_idx]
            job_val = row[job_idx]
            total_val = row[total_idx]
            if not date_val or not total_val:
                continue
            date_str = str(date_val).strip()
            if len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
                date = date_str
            else:
                try:
                    date = datetime.strptime(date_str, "%m/%d/%Y").strftime("%Y-%m-%d")
                except Exception:
                    continue
            customer = str(job_val or "").strip() or "Unknown"
            hours = round_hours(total_val)
            if hours <= 0:
                continue
            entries.append((date, employee, customer, hours, f"voice:{path.name}"))
    return entries


def parse_ocr_review(path: Path) -> List[Tuple[str, str, str, float, str]]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    if "Review" not in wb.sheetnames:
        return []
    ws = wb["Review"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    try:
        date_idx = header.index("date")
        emp_idx = header.index("employee")
        cust_idx = header.index("customer")
        hours_idx = header.index("hours")
        src_idx = header.index("source_image") if "source_image" in header else None
    except ValueError:
        return []
    entries: List[Tuple[str, str, str, float, str]] = []
    for row in rows[1:]:
        if not row or len(row) <= max(date_idx, emp_idx, cust_idx, hours_idx):
            continue
        date_val = row[date_idx]
        emp_val = row[emp_idx]
        cust_val = row[cust_idx]
        hours_val = row[hours_idx]
        if not date_val or not emp_val or not hours_val:
            continue
        date_str = str(date_val).strip()
        if len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
            date = date_str
        else:
            try:
                date = datetime.strptime(date_str, "%m/%d/%Y").strftime("%Y-%m-%d")
            except Exception:
                continue
        employee = str(emp_val).strip()
        customer = str(cust_val or "Unknown").strip() or "Unknown"
        hours = round_hours(hours_val)
        source = f"ocr:{row[src_idx]}" if src_idx is not None and row[src_idx] else "ocr"
        entries.append((date, employee, customer, hours, source))
    return entries


def parse_sheet_range(header_value: str) -> Optional[Tuple[str, str, str]]:
    # Example: "Boban Abbate  —  2/4/26 - 2/10/26"
    if not header_value:
        return None
    match = re.search(r"([0-9]{1,2}/[0-9]{1,2}/[0-9]{2})\s*-\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2})", str(header_value))
    if not match:
        return None
    start_raw, end_raw = match.groups()
    start = datetime.strptime(start_raw, "%m/%d/%y").strftime("%Y-%m-%d")
    end = datetime.strptime(end_raw, "%m/%d/%y").strftime("%Y-%m-%d")
    return start, end, f"{start_raw} - {end_raw}"


def date_label(date_str: str) -> str:
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    return f"{dt.strftime('%a')}-{dt.day}"


def clear_range(ws, start_row: int, end_row: int):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=7):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(40, max(10, max_len + 2))


def load_db_entries(db_path: str) -> Set[Tuple[str, str, str, float]]:
    """Load all approved entries from DB as a set of (date, employee_norm, customer_norm, hours)."""
    if not db_path or not Path(db_path).exists():
        return set()
    
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    # Join employees and customers to get names
    rows = c.execute("""
        SELECT te.work_date, e.name, c.name, te.hours
        FROM time_entries te
        JOIN employees e ON te.employee_id = e.id
        JOIN customers c ON te.customer_id = c.id
        WHERE te.status = 'APPROVED'
    """).fetchall()
    conn.close()
    
    entries = set()
    for date, emp, cust, hours in rows:
        entries.add((
            str(date),
            normalize_text(emp),
            normalize_text(cust),
            round_hours(hours)
        ))
    return entries


def main():
    args = parse_args()
    template = Path(args.template)
    out_path = Path(args.out)
    exports_root = Path(args.exports_root)
    
    # Load DB snapshot for drift detection
    db_snapshot = load_db_entries(args.db) if args.db else set()

    week_dirs = []
    if args.weeks:
        for name in [w.strip() for w in args.weeks.split(",") if w.strip()]:
            week_dirs.append(exports_root / name)
    else:
        week_dirs = sorted([p for p in exports_root.iterdir() if p.is_dir() and p.name.lower().startswith("week")])

    manual_entries: List[Tuple[str, str, str, float, str]] = []
    for week_dir in week_dirs:
        for path in sorted(week_dir.glob("*.xls*")):
            manual_entries.extend(parse_manual_entries(path))
    if args.manual_xlsx:
        for path in [p.strip() for p in args.manual_xlsx.split(",") if p.strip()]:
            manual_entries.extend(parse_voice_xlsx(Path(path)))
    if args.ocr_review:
        manual_entries.extend(parse_ocr_review(Path(args.ocr_review)))

    # Group entries by employee
    entries_by_employee: Dict[str, List[Tuple[str, str, str, float, str]]] = defaultdict(list)
    for date, employee, customer, hours, source in manual_entries:
        entries_by_employee[normalize_text(employee)].append((date, employee, customer, hours, source))

    wb = load_workbook(template, data_only=False)
    verify_fill = PatternFill("solid", fgColor="FFF2CC") # Yellow for OCR/Voice
    drift_fill = PatternFill("solid", fgColor="FFC7CE")  # Red for DB drift

    manual_sheet = wb.create_sheet("Manual Entries")
    manual_sheet.append(["date", "employee", "customer", "hours", "source"])
    for cell in manual_sheet[1]:
        cell.font = Font(bold=True)
    for row in sorted(manual_entries, key=lambda r: (r[0], r[1], r[2])):
        manual_sheet.append(list(row))
    autosize(manual_sheet)

    for name in wb.sheetnames:
        if name in ("Monthly Breakdown", "Manual Entries") or name.startswith("Week of"):
            continue
        ws = wb[name]
        header_val = ws["A1"].value
        range_info = parse_sheet_range(header_val)
        if not range_info:
            continue
        week_start, week_end, _label = range_info
        # employee name is stored on the sheet name
        employee_key = normalize_text(name)
        entries = [
            e for e in entries_by_employee.get(employee_key, [])
            if week_start <= e[0] <= week_end
        ]
        if not entries:
            continue

        # clear old values in rows 3-38 for A-G
        clear_range(ws, 3, 38)

        # Sort entries by date, then customer
        entries.sort(key=lambda r: (r[0], r[2]))
        row_idx = 3
        last_date = None
        for date, _emp, customer, hours, source in entries:
            if row_idx > 38:
                break
            label = date_label(date) if date != last_date else ""
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=customer)
            ws.cell(row=row_idx, column=6, value=hours)
            if source and source.startswith(("voice:", "ocr:")):
                ws.cell(row=row_idx, column=7, value=source)
            last_date = date

            # Check DB drift
            # We check if this exact entry exists in the DB snapshot
            is_in_db = False
            if db_snapshot:
                # Try exact match first
                key = (date, normalize_text(name), normalize_text(customer), round_hours(hours))
                if key in db_snapshot:
                    is_in_db = True
                else:
                    # Fallback: check if ANY entry for this emp/date matches hours (ignore customer alias drift?)
                    # For now, strict match on employee/date/hours/customer
                    pass

            # highlight verification rows
            fill_style = None
            if db_snapshot and not is_in_db:
                fill_style = drift_fill # Not in DB -> Red (Drift)
            elif source.startswith(("voice:", "ocr:")) or customer.lower() == "unknown":
                fill_style = verify_fill # Manual source -> Yellow (Review)

            if fill_style:
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).fill = fill_style

            row_idx += 1

    wb.save(out_path)
    print(f"Saved {out_path}")


if __name__ == "__main__":
    main()
