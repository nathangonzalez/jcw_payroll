#!/usr/bin/env python3
"""
Create a labor-only gap report: manual (exports + voice + OCR review) vs DB line items.
Manual-only entries indicate missing hours in DB.
"""

import argparse
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import fitz
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DATE_IN_NAME = re.compile(r"(\d{6})")
LINE_GROSS = re.compile(r"^\S+\s+([0-9]+\.?[0-9]*)\s+")
DAY_NAMES = {
    "mon",
    "monday",
    "tue",
    "tues",
    "tuesday",
    "wed",
    "wednesday",
    "thu",
    "thur",
    "thurs",
    "thursday",
    "fri",
    "friday",
    "sat",
    "saturday",
    "sun",
    "sunday",
}


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", required=True, help="Path to prod app.db")
    parser.add_argument("--pdfs", required=True, help="Comma-separated payroll PDF paths")
    parser.add_argument("--exports-root", required=True, help="Path to exports root")
    parser.add_argument("--weeks", default="", help="Comma-separated week folders (e.g. 'Week 2,Week 3')")
    parser.add_argument(
        "--manual-xlsx",
        default="",
        help="Optional comma-separated XLSX files with columns Date, Job, Task, Start, Lunch, End, Total.",
    )
    parser.add_argument(
        "--ocr-review",
        default="",
        help="Optional OCR review XLSX with Review sheet (date, employee, customer, hours).",
    )
    parser.add_argument("--dept", default="LABOR", help="Department label to include (default: LABOR)")
    parser.add_argument("--out", required=True, help="Output XLSX report")
    return parser.parse_args()


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def round_hours(value) -> float:
    try:
        return round(float(value), 2)
    except Exception:
        return 0.0


def parse_employee_from_filename(path: Path) -> str:
    name = path.stem
    name = re.sub(r"\s*\(\d+\)\s*$", "", name).strip()
    return name


def read_xls_rows(path: Path) -> List[List]:
    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_index(0)
    rows = []
    for r in range(sh.nrows):
        rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
    return rows


def read_xlsx_rows(path: Path) -> List[List]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def find_header_indices(rows: List[List]) -> Optional[Tuple[int, int, int, int]]:
    for idx, row in enumerate(rows[:10]):
        if not row:
            continue
        lowered = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        if "date" in lowered and "client name" in lowered and "hours per job" in lowered:
            return (idx, lowered.index("date"), lowered.index("client name"), lowered.index("hours per job"))
    return None


def is_day_name(value) -> bool:
    return isinstance(value, str) and value.strip().lower() in DAY_NAMES


def is_day_number(value) -> bool:
    return isinstance(value, (int, float)) and 1 <= int(value) <= 31


def build_date(year: int, month: int, day: int) -> str:
    return datetime(year, month, day).strftime("%Y-%m-%d")


def parse_manual_entries(path: Path, month_hint: Optional[str] = None) -> List[Tuple[str, str, str, float, str]]:
    rows = read_xlsx_rows(path) if path.suffix.lower() == ".xlsx" else read_xls_rows(path)
    header = find_header_indices(rows)
    if not header:
        return []
    header_row, date_col, client_col, hours_col = header
    file_mtime = datetime.utcfromtimestamp(path.stat().st_mtime)
    year = file_mtime.year
    month = file_mtime.month
    if month_hint:
        try:
            year, month = map(int, month_hint.split("-"))
        except Exception:
            pass

    entries = []
    pending = []
    current_date = None
    employee = parse_employee_from_filename(path)
    source = f"export:{path.name}"

    for row in rows[header_row + 1 :]:
        if not row or len(row) <= max(date_col, client_col, hours_col):
            continue
        date_cell = row[date_col]
        client_cell = row[client_col]
        hours_cell = row[hours_col]

        client = str(client_cell).strip() if client_cell is not None else ""
        hours = round_hours(hours_cell)

        if is_day_name(date_cell):
            if client and hours:
                pending.append((employee, client, hours))
            continue

        if is_day_number(date_cell):
            try:
                day_num = int(date_cell)
                use_year, use_month = year, month
                if not month_hint and day_num > file_mtime.day:
                    if use_month == 1:
                        use_year -= 1
                        use_month = 12
                    else:
                        use_month -= 1
                current_date = build_date(use_year, use_month, day_num)
            except ValueError:
                current_date = None
                pending = []
                continue
            if pending:
                for emp, cust, hrs in pending:
                    entries.append((current_date, emp, cust, hrs, source))
                pending = []
            if client and hours:
                entries.append((current_date, employee, client, hours, source))
            continue

        if current_date and client and hours:
            entries.append((current_date, employee, client, hours, source))

    return entries


def parse_voice_xlsx(path: Path) -> List[Tuple[str, str, str, float, str]]:
    wb = load_workbook(path, data_only=True)
    entries: List[Tuple[str, str, str, float, str]] = []
    for ws in wb.worksheets:
        employee = ws.title.replace("_", " ").strip()
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        try:
            date_idx = header.index("date")
            job_idx = header.index("job")
            total_idx = header.index("total")
        except ValueError:
            continue
        for row in rows[1:]:
            if not row or len(row) <= max(date_idx, job_idx, total_idx):
                continue
            date_val = row[date_idx]
            job_val = row[job_idx]
            total_val = row[total_idx]
            if not date_val or not total_val:
                continue
            date_str = str(date_val).strip()
            if len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
                date = date_str
            else:
                try:
                    date = datetime.strptime(date_str, "%m/%d/%Y").strftime("%Y-%m-%d")
                except Exception:
                    continue
            customer = str(job_val or "").strip() or "Unknown"
            hours = round_hours(total_val)
            if hours <= 0:
                continue
            entries.append((date, employee, customer, hours, f"voice:{path.name}"))
    return entries


def parse_ocr_review(path: Path) -> List[Tuple[str, str, str, float, str]]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    if "Review" not in wb.sheetnames:
        return []
    ws = wb["Review"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    try:
        date_idx = header.index("date")
        emp_idx = header.index("employee")
        cust_idx = header.index("customer")
        hours_idx = header.index("hours")
        src_idx = header.index("source_image") if "source_image" in header else None
    except ValueError:
        return []
    entries: List[Tuple[str, str, str, float, str]] = []
    for row in rows[1:]:
        if not row or len(row) <= max(date_idx, emp_idx, cust_idx, hours_idx):
            continue
        date_val = row[date_idx]
        emp_val = row[emp_idx]
        cust_val = row[cust_idx]
        hours_val = row[hours_idx]
        if not date_val or not emp_val or not hours_val:
            continue
        date_str = str(date_val).strip()
        if len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
            date = date_str
        else:
            try:
                date = datetime.strptime(date_str, "%m/%d/%Y").strftime("%Y-%m-%d")
            except Exception:
                continue
        employee = str(emp_val).strip()
        customer = str(cust_val or "Unknown").strip() or "Unknown"
        hours = round_hours(hours_val)
        source = f"ocr:{row[src_idx]}" if src_idx is not None and row[src_idx] else "ocr"
        entries.append((date, employee, customer, hours, source))
    return entries


def parse_employee_name(raw: str) -> str:
    raw = raw.strip()
    if "," in raw:
        last, first = [p.strip() for p in raw.split(",", 1)]
        return f"{first} {last}".strip()
    return raw


def extract_pdf_labor_employees(path: Path, dept_filter: str) -> List[str]:
    doc = fitz.open(path)
    employees = []
    pending_gross = None
    current_dept = None
    dept_filter = (dept_filter or "").strip().upper()
    for page in doc:
        text = page.get_text("text")
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            if line.startswith("* "):
                current_dept = line.upper()
                continue
            if line.startswith("** Total"):
                current_dept = None
            match = LINE_GROSS.match(line)
            if match:
                try:
                    pending_gross = float(match.group(1))
                except Exception:
                    pending_gross = None
                continue
            if pending_gross is not None and "," in line:
                if not dept_filter or (current_dept and dept_filter in current_dept):
                    emp = parse_employee_name(line)
                    employees.append(normalize_text(emp))
                pending_gross = None
    return employees


def week_range_from_filename(path: Path) -> Tuple[str, str]:
    match = DATE_IN_NAME.search(path.name)
    if not match:
        raise ValueError(f"Missing date in filename: {path.name}")
    mmddyy = match.group(1)
    week_end = datetime.strptime(mmddyy, "%m%d%y")
    week_start = week_end - timedelta(days=6)
    return week_start.strftime("%Y-%m-%d"), week_end.strftime("%Y-%m-%d")


def week_key_for_date(date_str: str, week_ranges: List[Tuple[str, str, str]]) -> Optional[str]:
    for week_id, start, end in week_ranges:
        if start <= date_str <= end:
            return week_id
    return None


def load_db_entries(conn, week_start: str, week_end: str) -> List[Tuple[str, str, str, float]]:
    cur = conn.cursor()
    cur.execute(
        """
        SELECT te.work_date, e.name, c.name, te.hours
        FROM time_entries te
        JOIN employees e ON e.id = te.employee_id
        JOIN customers c ON c.id = te.customer_id
        WHERE te.work_date >= ? AND te.work_date <= ?
        """,
        (week_start, week_end),
    )
    rows = cur.fetchall()
    return [(r[0], r[1], r[2], float(r[3])) for r in rows]


def to_counter(entries: List[Tuple[str, str, str, float, str]]) -> Counter:
    counter = Counter()
    for date, emp, cust, hours, _source in entries:
        key = (date, normalize_text(emp), normalize_text(cust), round_hours(hours))
        counter[key] += 1
    return counter


def expand_counter(counter: Counter, entries: List[Tuple[str, str, str, float, str]], label: str) -> List[List]:
    first_seen = {}
    sources = defaultdict(set)
    for date, emp, cust, hours, source in entries:
        key = (date, normalize_text(emp), normalize_text(cust), round_hours(hours))
        if key not in first_seen:
            first_seen[key] = (date, emp, cust, round_hours(hours))
        sources[key].add(source)
    rows = []
    for key, count in sorted(counter.items()):
        date, emp, cust, hours = first_seen.get(key, key)
        rows.append([label, date, emp, cust, hours, count, ",".join(sorted(sources.get(key, [])))])
    return rows


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(40, max(12, max_len + 2))


def main():
    args = parse_args()
    db_path = Path(args.db)
    exports_root = Path(args.exports_root)
    pdf_paths = [Path(p.strip()) for p in args.pdfs.split(",") if p.strip()]

    week_dirs = []
    if args.weeks:
        for name in [w.strip() for w in args.weeks.split(",") if w.strip()]:
            week_dirs.append(exports_root / name)
    else:
        week_dirs = sorted([p for p in exports_root.iterdir() if p.is_dir() and p.name.lower().startswith("week")])

    manual_entries: List[Tuple[str, str, str, float, str]] = []
    for week_dir in week_dirs:
        for path in sorted(week_dir.glob("*.xls*")):
            manual_entries.extend(parse_manual_entries(path))

    if args.manual_xlsx:
        for path in [p.strip() for p in args.manual_xlsx.split(",") if p.strip()]:
            manual_entries.extend(parse_voice_xlsx(Path(path)))

    if args.ocr_review:
        manual_entries.extend(parse_ocr_review(Path(args.ocr_review)))

    # Week ranges from PDFs + labor employee list
    week_ranges: List[Tuple[str, str, str]] = []
    labor_emps = set()
    for pdf in pdf_paths:
        week_start, week_end = week_range_from_filename(pdf)
        week_ranges.append((week_end, week_start, week_end))
        for emp in extract_pdf_labor_employees(pdf, args.dept):
            labor_emps.add(emp)

    # Filter manual to labor employees and known weeks
    manual_entries = [
        e
        for e in manual_entries
        if normalize_text(e[1]) in labor_emps and week_key_for_date(e[0], week_ranges)
    ]

    conn = sqlite3.connect(db_path)

    wb = Workbook()
    wb.remove(wb.active)
    red = PatternFill("solid", fgColor="FFD6D6")
    yellow = PatternFill("solid", fgColor="FFF2CC")

    summary = wb.create_sheet("Summary")
    summary.append(["week_end", "manual_only_count", "db_only_count"])
    for cell in summary[1]:
        cell.font = Font(bold=True)

    for week_end, week_start, week_end in sorted(week_ranges):
        db_entries = load_db_entries(conn, week_start, week_end)
        # Filter DB entries to labor employees
        db_entries = [
            (d, e, c, h)
            for d, e, c, h in db_entries
            if normalize_text(e) in labor_emps
        ]

        manual_week = [e for e in manual_entries if week_start <= e[0] <= week_end]
        manual_counter = to_counter(manual_week)
        db_counter = Counter()
        for date, emp, cust, hours in db_entries:
            key = (date, normalize_text(emp), normalize_text(cust), round_hours(hours))
            db_counter[key] += 1

        manual_only = manual_counter - db_counter
        db_only = db_counter - manual_counter

        ws = wb.create_sheet(f"Week_{week_end}")
        ws.append(["type", "date", "employee", "customer", "hours", "count", "source"])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        rows = []
        rows.extend(expand_counter(manual_only, manual_week, "manual_only"))
        rows.extend(expand_counter(db_only, [(d, e, c, h, "db") for d, e, c, h in db_entries], "db_only"))
        for row in rows:
            ws.append(row)
            fill = red if row[0] == "manual_only" else yellow
            for cell in ws[ws.max_row]:
                cell.fill = fill

        autosize(ws)
        summary.append([week_end, sum(manual_only.values()), sum(db_only.values())])

    autosize(summary)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved {out_path}")


if __name__ == "__main__":
    main()
