"""
Reconcile prod export vs PDF labor report (source of truth).

PDF Truth (LABOR dept gross wages):
  Week 1 (ending 2/3):  $7,823.75
  Week 2 (ending 2/10): $7,711.25
  Week 3 (ending 2/17): $8,032.50

This script:
  1. Reads the prod export XLSX
  2. Extracts per-employee hours from each weekly block
  3. Compares to PDF truth (derived hours)
  4. Generates a highlighted reconciliation XLSX
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
from collections import defaultdict

# ── PDF Truth ──────────────────────────────────────────────────────
# Per-employee gross wages from LABOR section of each PDF
PDF_LABOR = {
    "Week 1": {  # ending 2/3/26
        "Boban Abbate":    {"gross": 1700.00,  "rate": 42.50, "hours": 40.0},
        "Thomas Brinson":  {"gross": 1400.00,  "rate": 35.00, "hours": 40.0},
        "Jason Green":     {"gross": 1400.00,  "rate": 35.00, "hours": 40.0},
        "Phil Henderson":  {"gross": 1245.00,  "rate": 30.00, "hours": 41.5},
        "Doug Kinsey":     {"gross": 1278.75,  "rate": 30.00, "hours": 42.625},
        "Sean Matthew":    {"gross": 800.00,   "rate": 20.00, "hours": 40.0},
    },
    "Week 2": {  # ending 2/10/26
        "Boban Abbate":    {"gross": 1700.00,  "rate": 42.50, "hours": 40.0},
        "Thomas Brinson":  {"gross": 1400.00,  "rate": 35.00, "hours": 40.0},
        "Jason Green":     {"gross": 1400.00,  "rate": 35.00, "hours": 40.0},
        "Phil Henderson":  {"gross": 1200.00,  "rate": 30.00, "hours": 40.0},
        "Doug Kinsey":     {"gross": 1211.25,  "rate": 30.00, "hours": 40.375},
        "Sean Matthew":    {"gross": 800.00,   "rate": 20.00, "hours": 40.0},
    },
    "Week 3": {  # ending 2/17/26
        "Boban Abbate":    {"gross": 1700.00,  "rate": 42.50, "hours": 40.0},
        "Thomas Brinson":  {"gross": 1365.00,  "rate": 35.00, "hours": 39.0},
        "Jason Green":     {"gross": 1400.00,  "rate": 35.00, "hours": 40.0},
        "Phil Henderson":  {"gross": 1335.00,  "rate": 30.00, "hours": 44.5},
        "Doug Kinsey":     {"gross": 1402.50,  "rate": 30.00, "hours": 46.75},
        "Sean Matthew":    {"gross": 830.00,   "rate": 20.00, "hours": 41.5},
    },
}

# Employee rates (from prod export)
RATES = {
    "Boban Abbate": 42.50,
    "Thomas Brinson": 35.00,
    "Jason Green": 35.00,
    "Phil Henderson": 30.00,
    "Doug Kinsey": 30.00,
    "Sean Matthew": 20.00,
}

# Map week date ranges to week labels
WEEK_MAP = {
    "1/28": "Week 1",
    "2/4":  "Week 2",
    "2/11": "Week 3",
}

LABOR_EMPLOYEES = list(RATES.keys())

# ── Parse the prod export ──────────────────────────────────────────


def identify_week(header_text):
    """Given a header like 'Boban Abbate  —  1/28/26 - 2/3/26', return week label."""
    if not header_text:
        return None
    for date_prefix, week_label in WEEK_MAP.items():
        if date_prefix + "/" in header_text or date_prefix + "'" in header_text:
            return week_label
    # Try matching month/day pattern
    m = re.search(r'(\d+/\d+)/26\s*-', str(header_text))
    if m:
        start = m.group(1)
        for prefix, label in WEEK_MAP.items():
            if start == prefix:
                return label
    return None


def extract_employee_hours(ws, employee_name):
    """
    Extract hours per week from an employee sheet.
    Returns dict like {"Week 1": {"total_hours": 40, "entries": [...], "block_rows": (start, end)}, ...}
    """
    weeks = {}
    max_row = ws.max_row or 200

    # Find block headers by scanning column A + B for the employee name pattern
    # Headers are in row 1, 44, 87 approximately - they span the full row
    # Look for cells containing the date range pattern
    block_starts = []

    for row in range(1, max_row + 1):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            val = str(cell.value or "")
            if employee_name in val and "—" in val and "/26" in val:
                week = identify_week(val)
                if week:
                    block_starts.append((row, week))
                break

    # For each block, find the "Total:" row and sum hours from column F
    for i, (start_row, week_label) in enumerate(block_starts):
        # The data starts 2 rows after the header (header, then column labels, then data)
        data_start = start_row + 2
        
        # Find the end: look for "Total:" or "TOTAL:" in column E or F
        total_hours = 0.0
        entries = []
        total_row = None
        client_total_row = None

        for row in range(data_start, min(data_start + 40, max_row + 1)):
            # Check for TOTAL: marker in various columns
            for check_col in [1, 5, 6, 8, 9]:
                cv = str(ws.cell(row=row, column=check_col).value or "")
                if cv.strip().upper() == "TOTAL:":
                    total_row = row
                    break
                if cv.strip().upper().startswith("TOTAL"):
                    total_row = row
                    break
            if total_row:
                break

            # Read column F (Hours Per Job) - column 6
            hours_val = ws.cell(row=row, column=6).value
            date_val = ws.cell(row=row, column=1).value
            client_val = ws.cell(row=row, column=2).value

            if hours_val is not None:
                try:
                    h = float(hours_val)
                    if h > 0:
                        total_hours += h
                        entries.append({
                            "row": row,
                            "date": str(date_val or ""),
                            "client": str(client_val or ""),
                            "hours": h,
                        })
                except (ValueError, TypeError):
                    pass

        # Also read the right-side TOTAL (column J) which has SUM of client hours
        client_total = 0.0
        if total_row:
            jval = ws.cell(row=total_row, column=10).value  # Column J
            try:
                client_total = float(jval) if jval else 0.0
            except (ValueError, TypeError):
                client_total = 0.0

        # Also try reading the SUM(F) total
        f_total = 0.0
        if total_row:
            # Total row for column F is usually the row after TOTAL: or on a "Total:" row
            for search_row in range(total_row, min(total_row + 3, max_row + 1)):
                fval = ws.cell(row=search_row, column=6).value
                try:
                    f_total = float(fval) if fval else 0.0
                    if f_total > 0:
                        break
                except (ValueError, TypeError):
                    pass

        weeks[week_label] = {
            "computed_hours": round(total_hours, 4),
            "formula_total_j": client_total,
            "formula_total_f": f_total,
            "entries": entries,
            "block_start": start_row,
            "data_start": data_start,
            "total_row": total_row,
        }

    return weeks


def main():
    # Find the prod export
    export_path = None
    candidates = [
        os.path.expanduser("~/Downloads/Payroll_Breakdown_2026-02_1771623676826.xlsx"),
        "exports/2026-02/Payroll_Breakdown_2026-02.xlsx",
    ]
    for p in candidates:
        if os.path.exists(p):
            export_path = p
            break

    if not export_path:
        print("ERROR: Cannot find prod export XLSX. Tried:")
        for p in candidates:
            print(f"  {p}")
        return

    print(f"Reading prod export: {export_path}")

    # Read with data_only=True to get cached formula values
    wb = openpyxl.load_workbook(export_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    # Also read with formulas to see what's there
    wb_formulas = openpyxl.load_workbook(export_path, data_only=False)

    # ── Extract hours per employee per week ───────────────────────
    prod_data = {}
    for emp in LABOR_EMPLOYEES:
        if emp in wb.sheetnames:
            ws = wb[emp]
            weeks = extract_employee_hours(ws, emp)
            prod_data[emp] = weeks
            print(f"\n{emp}:")
            for wk, data in sorted(weeks.items()):
                print(f"  {wk}: computed={data['computed_hours']}h, "
                      f"J_total={data['formula_total_j']}h, "
                      f"F_total={data['formula_total_f']}h, "
                      f"entries={len(data['entries'])}")
        else:
            print(f"\n{emp}: SHEET NOT FOUND")
            prod_data[emp] = {}

    # ── Build reconciliation report ───────────────────────────────
    print("\n" + "=" * 80)
    print("RECONCILIATION: PROD EXPORT vs PDF LABOR REPORT (Source of Truth)")
    print("=" * 80)

    # Styles for the output workbook
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    bold_font = Font(bold=True, size=11)
    money_fmt = '#,##0.00'
    hrs_fmt = '#,##0.000'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    out_wb = openpyxl.Workbook()

    # ── Summary Sheet ─────────────────────────────────────────────
    ws_sum = out_wb.active
    ws_sum.title = "Reconciliation Summary"

    headers = ["Employee", "Rate",
               "Wk1 PDF Hrs", "Wk1 Prod Hrs", "Wk1 Δ Hrs", "Wk1 PDF $", "Wk1 Prod $", "Wk1 Δ $",
               "Wk2 PDF Hrs", "Wk2 Prod Hrs", "Wk2 Δ Hrs", "Wk2 PDF $", "Wk2 Prod $", "Wk2 Δ $",
               "Wk3 PDF Hrs", "Wk3 Prod Hrs", "Wk3 Δ Hrs", "Wk3 PDF $", "Wk3 Prod $", "Wk3 Δ $"]

    for col, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    row = 2
    all_gaps = []

    for emp in LABOR_EMPLOYEES:
        rate = RATES[emp]
        ws_sum.cell(row=row, column=1, value=emp).font = bold_font
        ws_sum.cell(row=row, column=2, value=rate).number_format = money_fmt

        for wi, wk_label in enumerate(["Week 1", "Week 2", "Week 3"]):
            col_offset = 3 + wi * 6  # columns for this week

            pdf = PDF_LABOR[wk_label].get(emp, {})
            pdf_hrs = pdf.get("hours", 0)
            pdf_gross = pdf.get("gross", 0)

            prod_wk = prod_data.get(emp, {}).get(wk_label, {})
            prod_hrs = prod_wk.get("computed_hours", 0)
            # Use formula total if computed is 0
            if prod_hrs == 0 and prod_wk.get("formula_total_j", 0) > 0:
                prod_hrs = prod_wk["formula_total_j"]
            if prod_hrs == 0 and prod_wk.get("formula_total_f", 0) > 0:
                prod_hrs = prod_wk["formula_total_f"]

            prod_gross = prod_hrs * rate
            delta_hrs = prod_hrs - pdf_hrs
            delta_gross = prod_gross - pdf_gross

            c_pdf_hrs = ws_sum.cell(row=row, column=col_offset, value=pdf_hrs)
            c_prod_hrs = ws_sum.cell(row=row, column=col_offset + 1, value=prod_hrs)
            c_delta_hrs = ws_sum.cell(row=row, column=col_offset + 2, value=delta_hrs)
            c_pdf_gross = ws_sum.cell(row=row, column=col_offset + 3, value=pdf_gross)
            c_prod_gross = ws_sum.cell(row=row, column=col_offset + 4, value=prod_gross)
            c_delta_gross = ws_sum.cell(row=row, column=col_offset + 5, value=delta_gross)

            c_pdf_hrs.number_format = hrs_fmt
            c_prod_hrs.number_format = hrs_fmt
            c_delta_hrs.number_format = hrs_fmt
            c_pdf_gross.number_format = money_fmt
            c_prod_gross.number_format = money_fmt
            c_delta_gross.number_format = money_fmt

            for c in [c_pdf_hrs, c_prod_hrs, c_delta_hrs, c_pdf_gross, c_prod_gross, c_delta_gross]:
                c.border = thin_border

            # Highlight
            if abs(delta_hrs) < 0.01:
                c_delta_hrs.fill = green_fill
                c_delta_gross.fill = green_fill
            else:
                c_delta_hrs.fill = red_fill
                c_delta_gross.fill = red_fill
                c_prod_hrs.fill = yellow_fill
                c_prod_gross.fill = yellow_fill
                all_gaps.append({
                    "employee": emp,
                    "week": wk_label,
                    "pdf_hrs": pdf_hrs,
                    "prod_hrs": prod_hrs,
                    "delta_hrs": delta_hrs,
                    "delta_gross": delta_gross,
                    "prod_entries": prod_wk.get("entries", []),
                })

        row += 1

    # Totals row
    row += 1
    ws_sum.cell(row=row, column=1, value="TOTALS").font = bold_font
    for wi, wk_label in enumerate(["Week 1", "Week 2", "Week 3"]):
        col_offset = 3 + wi * 6
        pdf_total = sum(v["gross"] for v in PDF_LABOR[wk_label].values())
        prod_total = 0
        for emp in LABOR_EMPLOYEES:
            rate = RATES[emp]
            prod_wk = prod_data.get(emp, {}).get(wk_label, {})
            prod_hrs = prod_wk.get("computed_hours", 0)
            if prod_hrs == 0 and prod_wk.get("formula_total_j", 0) > 0:
                prod_hrs = prod_wk["formula_total_j"]
            if prod_hrs == 0 and prod_wk.get("formula_total_f", 0) > 0:
                prod_hrs = prod_wk["formula_total_f"]
            prod_total += prod_hrs * rate

        c = ws_sum.cell(row=row, column=col_offset + 3, value=pdf_total)
        c.number_format = money_fmt
        c.font = bold_font
        c = ws_sum.cell(row=row, column=col_offset + 4, value=prod_total)
        c.number_format = money_fmt
        c.font = bold_font
        c = ws_sum.cell(row=row, column=col_offset + 5, value=prod_total - pdf_total)
        c.number_format = money_fmt
        c.font = bold_font
        if abs(prod_total - pdf_total) < 0.01:
            c.fill = green_fill
        else:
            c.fill = red_fill

    # Column widths
    ws_sum.column_dimensions['A'].width = 18
    ws_sum.column_dimensions['B'].width = 8
    for col in range(3, 21):
        ws_sum.column_dimensions[get_column_letter(col)].width = 12

    # ── Fixes Needed Sheet ────────────────────────────────────────
    ws_fix = out_wb.create_sheet("Fixes Needed")
    fix_headers = ["Employee", "Week", "PDF Hours", "Prod Hours", "Δ Hours", "Δ Gross",
                   "Action Needed", "Prod Entries (current)"]
    for col, h in enumerate(fix_headers, 1):
        cell = ws_fix.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    fix_row = 2
    for gap in all_gaps:
        ws_fix.cell(row=fix_row, column=1, value=gap["employee"]).font = bold_font
        ws_fix.cell(row=fix_row, column=2, value=gap["week"])
        ws_fix.cell(row=fix_row, column=3, value=gap["pdf_hrs"]).number_format = hrs_fmt
        ws_fix.cell(row=fix_row, column=4, value=gap["prod_hrs"]).number_format = hrs_fmt

        c = ws_fix.cell(row=fix_row, column=5, value=gap["delta_hrs"])
        c.number_format = hrs_fmt
        c.fill = red_fill

        c = ws_fix.cell(row=fix_row, column=6, value=gap["delta_gross"])
        c.number_format = money_fmt
        c.fill = red_fill

        if gap["delta_hrs"] < 0:
            action = f"ADD {abs(gap['delta_hrs'])} hours to prod"
        else:
            action = f"REMOVE {abs(gap['delta_hrs'])} hours from prod"
        ws_fix.cell(row=fix_row, column=7, value=action).fill = yellow_fill

        # List current entries
        entries_str = "; ".join(
            f"{e['date']} {e['client']} {e['hours']}h"
            for e in gap["prod_entries"]
        )
        ws_fix.cell(row=fix_row, column=8, value=entries_str[:500])

        for col in range(1, 9):
            ws_fix.cell(row=fix_row, column=col).border = thin_border

        fix_row += 1

    ws_fix.column_dimensions['A'].width = 18
    ws_fix.column_dimensions['B'].width = 10
    ws_fix.column_dimensions['G'].width = 30
    ws_fix.column_dimensions['H'].width = 80

    # ── Employee Detail Sheets (entries per week) ─────────────────
    for emp in LABOR_EMPLOYEES:
        ws_det = out_wb.create_sheet(emp[:20])
        det_headers = ["Week", "Date", "Client", "Hours", "Source"]
        for col, h in enumerate(det_headers, 1):
            cell = ws_det.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font

        det_row = 2
        for wk_label in ["Week 1", "Week 2", "Week 3"]:
            prod_wk = prod_data.get(emp, {}).get(wk_label, {})
            pdf = PDF_LABOR[wk_label].get(emp, {})

            for entry in prod_wk.get("entries", []):
                ws_det.cell(row=det_row, column=1, value=wk_label)
                ws_det.cell(row=det_row, column=2, value=entry["date"])
                ws_det.cell(row=det_row, column=3, value=entry["client"])
                ws_det.cell(row=det_row, column=4, value=entry["hours"])
                ws_det.cell(row=det_row, column=5, value="prod")
                det_row += 1

            # Summary row for this week
            prod_hrs = prod_wk.get("computed_hours", 0)
            pdf_hrs = pdf.get("hours", 0)
            ws_det.cell(row=det_row, column=1, value=wk_label)
            ws_det.cell(row=det_row, column=2, value="TOTAL")
            ws_det.cell(row=det_row, column=4, value=prod_hrs).font = bold_font
            ws_det.cell(row=det_row, column=5,
                        value=f"PDF={pdf_hrs}h, Δ={round(prod_hrs - pdf_hrs, 3)}h")
            if abs(prod_hrs - pdf_hrs) > 0.01:
                ws_det.cell(row=det_row, column=4).fill = red_fill
                ws_det.cell(row=det_row, column=5).fill = red_fill
            else:
                ws_det.cell(row=det_row, column=4).fill = green_fill
            det_row += 1
            det_row += 1  # blank row

    # ── Save ──────────────────────────────────────────────────────
    out_path = "exports/reconcile/2026-02/labor_reconciliation_report.xlsx"
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    out_wb.save(out_path)
    print(f"\n✅ Report saved: {out_path}")

    # ── Console summary ───────────────────────────────────────────
    print("\n" + "=" * 80)
    print("GAP SUMMARY")
    print("=" * 80)
    if not all_gaps:
        print("✅ All weeks match! No gaps found.")
    else:
        for gap in all_gaps:
            print(f"  ❌ {gap['employee']} {gap['week']}: "
                  f"PDF={gap['pdf_hrs']}h, Prod={gap['prod_hrs']}h, "
                  f"Δ={gap['delta_hrs']:+.3f}h (${gap['delta_gross']:+.2f})")

    print("\n── WEEK TOTALS ──")
    for wk_label in ["Week 1", "Week 2", "Week 3"]:
        pdf_total = sum(v["gross"] for v in PDF_LABOR[wk_label].values())
        prod_total = 0
        for emp in LABOR_EMPLOYEES:
            rate = RATES[emp]
            prod_wk = prod_data.get(emp, {}).get(wk_label, {})
            prod_hrs = prod_wk.get("computed_hours", 0)
            if prod_hrs == 0 and prod_wk.get("formula_total_j", 0) > 0:
                prod_hrs = prod_wk["formula_total_j"]
            if prod_hrs == 0 and prod_wk.get("formula_total_f", 0) > 0:
                prod_hrs = prod_wk["formula_total_f"]
            prod_total += prod_hrs * rate

        delta = prod_total - pdf_total
        status = "✅" if abs(delta) < 0.01 else "❌"
        print(f"  {status} {wk_label}: PDF=${pdf_total:,.2f}, Prod=${prod_total:,.2f}, Δ=${delta:+,.2f}")


if __name__ == "__main__":
    main()