import openpyxl
wb = openpyxl.load_workbook('exports/reconcile/2026-02/post_patch_export.xlsx', data_only=True)

ws = wb['Phil Henderson']
print(f'Phil Henderson: {ws.max_row} rows, {ws.max_column} cols')
sep = "  "
for r in range(1, min(ws.max_row+1, 150)):
    vals = []
    for c in range(1, min(ws.max_column+1, 12)):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            vals.append(f'[{c}]{v}')
    if vals:
        line = sep.join(vals)
        print(f'  R{r}: {line}'[:120])